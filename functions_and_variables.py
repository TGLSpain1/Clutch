# LIBRARY TO WORK WITH JSON FORMAT FILES
import json

#LIBRARY TO BE ABLE TO CHECK IF EXCEL EXISTS AND IF NOT CREATE IT
import os

# LIBRARY TO BE ABLE TO PAUSE SCRIPT EXECUTION
import time

# LIBRARY TO CREATE PSEUDO RANDOM INTEGER VALUES. USED WITH TIME LIBRARY TO GENERATE RANDOM PAUSES BETWEEN CALLS AND SIMULATE MORE HUMAN BEHAVIOR
import random

# LIBRARY WHICH WORKS LIKE SELENIUM BUT MUCH HARDER TO DETECT BY CLOUDFLARE
import undetected_chromedriver as uc

# MODULE TO BE ABLE TO FIND WEB ELEMENTS
from selenium.webdriver.common.by import By

# MODULE TO BE ABLE TO HANDLE THE ELEMENT NOT FOUND EXCEPTION
from selenium.common.exceptions import NoSuchElementException

# LIBRARY TO BE ABLE TO SAVE TO EXCEL FILE
import openpyxl
from openpyxl import load_workbook

# LIBRARY TO BE ABLE TO PRINT WEIRD CHARACTERS IN CONSOLE
from unidecode import unidecode

chrome_driver = "C:\\Users\\tglhi\\Desktop\\Estadisticas\\Python Stuff\\Clutch\\chromedriver.exe"
base_url = "https://clutch.co/us/call-centers"
company_name_for_email = "My Company"
email_text = "Hello,\n\nI hope you are well"

# RANDOM PAUSE IN SECONDS SO SCRIPT BEHAVIOUR LOOKS MORE HUMAN
def random_pause(lowest, highest):
	print("random wait")
	time.sleep(random.randint(lowest,highest))

## TRACKING VARIABLE FUNCTIONS
# READ VARIABLE FROM JSON FILE
def read_from_json(filename, my_var):
	with open(filename, "r") as f:
		read_dict = json.load(f)
		read_var = read_dict[my_var]
		return read_var

# WRITE TO JSON FILE
def write_to_json(my_dict, filename):
	json_object = json.dumps(my_dict)
	with open(filename, "w") as f:
		f.write(json_object)

# UPDATE TRACKED VARIABLES
def update_tracked_variables(total_company_number, results_page_number, results_page_company):
	print("updating tracked variables")
	dict_to_update_tracked_variables = {"total_company_number": total_company_number, "results_page_number": results_page_number, "results_page_company": results_page_company}
	write_to_json(dict_to_update_tracked_variables, "tracking_variables.json")

# MAKE SURE THESE TWO VARIABLES ARE SET TO 0 IN THE JSON FILE IF YOU'RE RUNNING THE SCRIPT FOR THE FIRST TIME. 
# IF YOU'RE RUNNING IT BECAUSE IT HAD PREVIOUSLY STOPPED BEFORE, KEEP THE VALUES SAVED FOR THE VARIABLES
def set_total_company_number():
	total_company_number = read_from_json("tracking_variables.json", "total_company_number")
	return total_company_number

def set_results_page_number():
	results_page_number = read_from_json("tracking_variables.json", "results_page_number")
	return results_page_number

def set_results_page_company():
	results_page_company = read_from_json("tracking_variables.json", "results_page_company")
	return results_page_company

## DRIVER SET UP FUNCTIONS
def set_up_driver():
	options = uc.ChromeOptions()
	options.debugger_address = "127.0.0.1:9222"
	driver = uc.Chrome(executable_path = chrome_driver, options=options)
	return driver

## FUNCTIONS FOR GENERATING ALL RESULTS PAGES URLS
# GET NUMBER OF RESULT PAGES
def get_number_of_results_pages(driver):
	print("Getting number of results pages")
	page_navigation = driver.find_element(By.XPATH, "//nav[@aria-label='Page navigation']")
	last_page_item = page_navigation.find_element(By.CLASS_NAME, "page-item.last")
	last_page_href = last_page_item.find_element(By.TAG_NAME, "a")
	number_of_pages = int(last_page_href.get_attribute("data-page"))
	return number_of_pages

# CREATE URLS FOR ALL RESULT PAGES
def create_results_pages_urls(number_of_pages):
	print("Creating results_pages_urls")
	results_pages_urls = [base_url]
	for i in range(1, (number_of_pages + 1)):
		results_page_url = base_url + f"?page={i}"
		results_pages_urls.append(results_page_url)
	return results_pages_urls

# NAVIGATE THROUGH ALL RESULT PAGES
def navigate_results_page(results_pages_url, driver):
	print(f"navigating to {results_pages_url}")
	driver.get(results_pages_url)
	random_pause(1,3)

# CHECK IF EXCEL FILE EXISTS AND IF NOT CREATE IT
def check_excel_exists():
	print("checking if sent_companies excel exists")
	if not os.path.exists("sent_companies.xlsx"):    
		sent_companies_empty_wb = openpyxl.Workbook()
		sent_companies_empty_ws = sent_companies_empty_wb.active
		data = (("name", "website", "location", "employees", "services", "contact_url", "email_sent"))
		sent_companies_empty_ws.append(data)
		sent_companies_empty_wb.save("sent_companies.xlsx")
		print("sent_companies excel created")
	else: 
		print("Excel already exists. No need to create it again")

## FUNCTIONS FOR GETTING ALL THE INFORMATION FOR EACH COMPANY IN COMPANY RESULTS PAGE (NAME, LOCATION, EMPLOYEES, WEBSITE...)
# GET COMPANY BLOCKS
def get_company_blocks(driver):
	print("Getting company blocks")
	companies_table = driver.find_element(By.CLASS_NAME, "directory-list")
	companies_li_elements = companies_table.find_elements(By.CLASS_NAME, "provider.provider-row")
	return companies_li_elements

# GET NAME OF COMPANY
def get_company_name(company_li_element):
	name_element = company_li_element.find_element(By.TAG_NAME, "h3")
	name = unidecode(name_element.text)
	return name

# GET WEBSITE OF COMPANY
def get_company_website(company_li_element):
	right_nav_bar = company_li_element.find_element(By.CLASS_NAME, "nav-right-profile")
	website_url_element = right_nav_bar.find_element(By.CLASS_NAME, "website-link.website-link-a")
	website_url = website_url_element.find_element(By.TAG_NAME, "a").get_attribute("href").split("?")[0]
	return website_url

# GET LOCATION OF COMPANY
def get_company_location(company_li_element):
	company_info_element = company_li_element.find_element(By.CLASS_NAME, "col-md-3.provider-info__details")
	location_element = company_info_element.find_element(By.CSS_SELECTOR, "div[data-content='<i>Location</i>']")
	location = location_element.find_element(By.TAG_NAME, "span").text
	return location

# GET NUMBER OF EMPLOYEES
def get_employees_number(company_li_element):	
	company_info_element = company_li_element.find_element(By.CLASS_NAME, "col-md-3.provider-info__details")
	employees_element = company_info_element.find_element(By.CSS_SELECTOR, "div[data-content='<i>Employees</i>']")
	employees = employees_element.find_element(By.TAG_NAME, "span").text
	return employees

# GET SERVICE FOCUS
def get_service_focus(company_li_element):	
	company_graph = company_li_element.find_element(By.CSS_SELECTOR, "div[class='carousel-inner']")
	company_graph_data_items = company_graph.find_elements(By.CLASS_NAME, "carousel-item")
	company_service_focus = ""
	for company_graph_data_item in company_graph_data_items:
		company_service_focus += (", " + company_graph_data_item.get_attribute("innerText"))
	company_service_focus = company_service_focus[2:]
	return company_service_focus

# GET CONTACT URL
def get_contact_url(company_li_element):
	contact_details_element = company_li_element.find_element(By.CLASS_NAME, "provider-detail.col-md-2")
	website_contact_element = contact_details_element.find_element(By.CLASS_NAME, "website-contact")
	website_contact_url_element = website_contact_element.find_element(By.TAG_NAME, "a")
	website_contact_url = website_contact_url_element.get_attribute("href")
	return website_contact_url

# GET ALL COMPANY INFO
def get_all_company_info(company_li_element):
	company_dict = {}	
	company_dict["name"] = get_company_name(company_li_element)
	print(f"Getting all company info for {company_dict['name']}")
	company_dict["website"] = get_company_website(company_li_element)
	company_dict["location"] = get_company_location(company_li_element)
	company_dict["elements"] = get_employees_number(company_li_element)
	company_dict["services"] = get_service_focus(company_li_element)
	company_dict["contact_url"] = get_contact_url(company_li_element)
	return company_dict

# CHECKS IF TEXT FIELD IS EMPTY AND IF NOT CLEARS IT
def check_text_field_empty(element):
	print("Checkin if text field is empty")
	if element.get_attribute("value") != "":
		element.clear()

## FUNCTIONS FOR FILLING OUT CONTACT FORM AND SENDING EMAIL
# OPENING CONTACT PAGE FOR COMPANY AND WAITING FOR SUBMIT BUTTON TO LOAD TO MAKE SURE SCRIPT DOESN'T GO TOO FAST
def wait_for_submit(unsent_company, driver):
	print("Waiting for submit button")		
	submit_button = driver.find_element(By.ID, "submit")
	print("submit button found")
	return submit_button

# FIND AND FILL COMPANY FIELD
def fill_company(driver):
	print("Writing company name")
	company_name_input = driver.find_element(By.ID, "company")
	check_text_field_empty(company_name_input)
	company_name_input.send_keys(company_name_for_email)

# FIND AND SELECT PARTNERSHIP BUTTON
def select_partnership(driver):
	print("Clicking partnership button")
	partnership_button = driver.find_element(By.ID, "partnership")
	if partnership_button.is_selected() == False:
		partnership_button.click()

# INTRODUCE EMAIL TEXT
def write_email(driver, email_text):
	print("Writing email")
	email_message = email_text
	email_text = driver.find_element(By.TAG_NAME, "textarea")
	check_text_field_empty(email_text)
	email_text.send_keys(email_message)

# FIND AND DESELECT EMAIL AND SHORTLIST CHECKBOXES
def deselect_email_and_shortlist(driver):
	print("Deselecting email checkbox")
	send_own_email_box = driver.find_element(By.ID, "checkbox")
	if send_own_email_box.is_selected():
		send_own_email_box.click()		
	print("Deselecting shortlist checkbox")
	add_to_shortlist_box = driver.find_element(By.ID, "shortlist_checkbox")
	if add_to_shortlist_box.is_selected():
		add_to_shortlist_box.click()

# CLICK ON SEND BUTTON AND VERIFY EMAIL IS CORRECTLY SENT
def send_email(driver, submit_button):
	print("Sending email")
	submit_button.click()
	print("Send email button clicked")
	time.sleep(1)
	driver.find_element(By.CLASS_NAME, "redirect")
	print("Email succesfully sent")

# SET EMAIL SENT TO YES AND SAVE TO EXCEL
def update_email_sent_excel(unsent_company):
	print("Updating email_sent for company")
	unsent_company["email_sent"] = "Yes"
	print(unsent_company)
	sent_companies_wb = load_workbook(filename = "sent_companies.xlsx")
	sent_companies_ws = sent_companies_wb.active
	my_data = ((list(unsent_company.values())))
	sent_companies_ws.append(my_data)
	sent_companies_wb.save("sent_companies.xlsx")

# FILL OUT EMAIL FORM AND SEND EMAIL
def send_email_process(unsent_company, driver, email_text):
	tries = 0
	print("Starting email sending process")
	if tries <= 5:
		try:		
			driver.get(unsent_company["contact_url"])
			fill_company(driver)
			select_partnership(driver)
			write_email(driver, email_text)
			deselect_email_and_shortlist(driver)
			submit_button = wait_for_submit(unsent_company, driver)
			send_email(driver, submit_button)
			update_email_sent_excel(unsent_company)
		except NoSuchElementException:
			tries += 1
			print(f"Element not found. Retrying attempt {tries}")
			driver.get(unsent_company["contact_url"])
			fill_company(driver)
			select_partnership(driver)
			write_email(driver, email_text)
			deselect_email_and_shortlist(driver)
			submit_button = wait_for_submit(unsent_company, driver)
			# send_email(driver, submit_button)
			update_email_sent_excel(unsent_company)




